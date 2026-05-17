"""
handlers/xlsx.py — XLSX ingestion handler.

Contract: 90 System/_ingestion_contract.md §1, §2, §5, §9
HR-02 deliverable (S03).

Uses openpyxl (read_only=False for formula access) to extract spreadsheet
content. Per-sheet output structure:

    ## Sheet: <name>

    <markdown table of data rows, capped at 1000 visible rows>

    ... N more rows in original at [[99 Workspace/_originals/<hash>_<name>]]

    ## Formulas

    <cell_ref>: <formula>  (one per line, every formula cell on this sheet)

Header heuristic:
  - If the first non-empty row has all string values → treat as header row.
  - Otherwise → synthesise headers as "A", "B", "C", … (column letters).

Merged cells: annotated as `[merged: <range>]` in the header row or
              wherever the merge starts.

Type written to frontmatter: 'spreadsheet' (contract §2).

Formulas: openpyxl reads the formula string (e.g. '=SUM(A1:A10)').
  - The main table renders the formula string verbatim (no evaluation).
  - The `## Formulas` section at end of each sheet lists every formula cell.
  - NOTE: data_only=False is required; data_only=True loses formulas.
    Computed values are available ONLY in cached_value if Excel saved them.
    We show the formula in both places — the contract says "computed value"
    but openpyxl without Excel cannot evaluate; we show cached_value when
    available, falling back to the formula string.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from .text import HandlerResult, _PROVENANCE_VALUE

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

ROW_CAP = 1000  # maximum data rows rendered per sheet (excluding header)


def _openpyxl_version() -> str:
    try:
        import importlib.metadata
        return importlib.metadata.version("openpyxl")
    except Exception:
        return "unknown"


# ──────────────────────────────────────────────────────────────────────────────
# Cell value helpers
# ──────────────────────────────────────────────────────────────────────────────

def _cell_display(cell) -> str:
    """Return the display string for a cell.

    For formula cells:
    1. Try the cached value (Excel-computed, stored in the xlsx file).
    2. Fall back to the formula string itself.

    For all other cells, str() the value.
    """
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, str) and val.startswith("="):
        # It's a formula. Try cached_value first.
        cached = getattr(cell, "cached_value", None)
        if cached is not None:
            return str(cached)
        return val  # show the formula string if no cached value
    if isinstance(val, float):
        # Avoid unnecessary .0 for whole numbers
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


def _is_formula(cell) -> bool:
    val = cell.value
    return isinstance(val, str) and val.startswith("=")


# ──────────────────────────────────────────────────────────────────────────────
# Header heuristic
# ──────────────────────────────────────────────────────────────────────────────

def _infer_header(first_data_row: list) -> tuple[list[str], bool]:
    """Decide whether the first data row is a header row.

    Returns (headers, is_real_header):
    - is_real_header=True  → first_data_row IS the header; skip it in body.
    - is_real_header=False → synthesised column letters; use first_data_row as data.

    Heuristic: header if ALL non-empty cells in the first row are strings
    (not formulas, not numbers, not dates).
    """
    if not first_data_row:
        return [], False

    non_empty = [v for v in first_data_row if v is not None and str(v).strip()]
    if not non_empty:
        return [get_column_letter(i + 1) for i in range(len(first_data_row))], False

    all_strings = all(isinstance(v, str) and not str(v).startswith("=") for v in non_empty)
    if all_strings:
        # Real header — sanitise pipe chars
        return [str(v).replace("|", "\\|") for v in first_data_row], True

    # Synthesised headers
    return [get_column_letter(i + 1) for i in range(len(first_data_row))], False


# ──────────────────────────────────────────────────────────────────────────────
# Merged-cell tracking
# ──────────────────────────────────────────────────────────────────────────────

def _merged_cell_map(ws) -> dict[tuple[int, int], str]:
    """Map (row, col) → merge-range string for all merged cells on the sheet.

    Only the top-left cell of each merge range is mapped (it has the value).
    Other cells in the range are mapped to None (empty).

    NOTE (2026-05-16 fix): openpyxl >=3.1 returns ``(row_idx, col_idx)`` tuples
    from ``MergedCellRange.cells`` (and from ``.rows``), not Cell objects.
    The prior implementation iterated ``.rows`` and accessed ``cell.row`` /
    ``cell.column`` on what was actually a tuple, raising
    ``AttributeError: 'tuple' object has no attribute 'row'`` on every
    workbook that contains a merged region. ``.cells`` is the documented
    coordinate iterator — use it directly.
    """
    merged: dict[tuple[int, int], str] = {}
    for merge_range in ws.merged_cells.ranges:
        rng_str = str(merge_range)
        for row_idx, col_idx in merge_range.cells:
            merged[(row_idx, col_idx)] = rng_str
    return merged


# ──────────────────────────────────────────────────────────────────────────────
# Markdown table builder
# ──────────────────────────────────────────────────────────────────────────────

def _build_md_table(
    headers: list[str],
    rows: list[list[str]],
    merged_map: dict[tuple[int, int], str],
    data_start_row: int,  # 1-based sheet row index of first data row
) -> list[str]:
    """Render headers + data rows as a markdown pipe table.

    Merged cells in the header row get '[merged: <range>]' appended.
    Other merged cells show their value normally (openpyxl exposes the
    top-left value; non-top-left cells are None / empty).
    """
    if not headers:
        return []

    ncols = len(headers)

    # Annotate header cells that are merged
    annotated_headers: list[str] = []
    for col_idx, h in enumerate(headers, start=1):
        rng = merged_map.get((data_start_row - 1, col_idx))  # header is one row above data start
        if rng:
            annotated_headers.append(f"{h} [merged: {rng}]")
        else:
            annotated_headers.append(h)

    # Pad rows
    padded_rows = [r + [""] * (ncols - len(r)) for r in rows]

    # Compute column widths
    all_rows = [annotated_headers] + padded_rows
    widths = [
        max(len(str(all_rows[ri][ci])) for ri in range(len(all_rows)))
        for ci in range(ncols)
    ]
    widths = [max(w, 3) for w in widths]

    def fmt_row(cells: list[str]) -> str:
        return "| " + " | ".join(
            str(c).ljust(widths[i]) for i, c in enumerate(cells)
        ) + " |"

    lines: list[str] = []
    lines.append(fmt_row(annotated_headers))
    lines.append("| " + " | ".join("-" * widths[j] for j in range(ncols)) + " |")
    for row in padded_rows:
        lines.append(fmt_row(row))
    return lines


# ──────────────────────────────────────────────────────────────────────────────
# Per-sheet extractor
# ──────────────────────────────────────────────────────────────────────────────

def _extract_sheet(
    ws,
    original_ref: str,
) -> tuple[list[str], list[str], int, int, int]:
    """Extract one worksheet to markdown.

    Returns
    -------
    body_lines : list[str]
        Markdown lines for the sheet section (header + table + overflow footer).
    formula_lines : list[str]
        Lines for the ## Formulas section (empty if no formulas).
    row_count : int
        Total data rows (excluding header), regardless of cap.
    formula_count : int
        Number of formula cells found.
    table_count : int
        Always 1 (each sheet is one table) — 0 if sheet is empty.
    """
    body_lines: list[str] = []
    formula_lines: list[str] = []
    formula_count = 0

    # Collect all rows (list of lists of raw cell objects)
    raw_rows: list[list] = []
    for row in ws.iter_rows():
        raw_rows.append(list(row))

    # Remove trailing completely-empty rows
    while raw_rows and all(c.value is None for c in raw_rows[-1]):
        raw_rows.pop()

    if not raw_rows:
        return body_lines, formula_lines, 0, 0, 0

    # Determine number of columns (max non-empty column index + 1)
    ncols = max(
        (len(row) for row in raw_rows),
        default=0
    )

    # First row values (raw)
    first_row_values = [c.value for c in raw_rows[0]]
    headers, is_real_header = _infer_header(first_row_values)

    # Merged cell map
    merged_map = _merged_cell_map(ws)

    # Data rows start index (0-based into raw_rows)
    data_start_idx = 1 if is_real_header else 0
    data_rows_raw = raw_rows[data_start_idx:]
    total_data_rows = len(data_rows_raw)

    # Cap display rows
    display_rows_raw = data_rows_raw[:ROW_CAP]
    overflow = total_data_rows - len(display_rows_raw)

    # Build display rows (list of list of str)
    display_rows: list[list[str]] = []
    for row in display_rows_raw:
        row_cells: list[str] = []
        for cell in row:
            row_cells.append(_cell_display(cell))
        display_rows.append(row_cells[:ncols])

    # Build header row (sheet row number for merged_map lookup)
    # Header is row 1 of the sheet (1-based)
    header_sheet_row = 1
    data_sheet_row_start = header_sheet_row + (1 if is_real_header else 0)

    md_table = _build_md_table(
        headers,
        display_rows,
        merged_map,
        data_start_row=data_sheet_row_start,
    )
    body_lines.extend(md_table)

    # Overflow footer
    if overflow > 0:
        body_lines.append("")
        body_lines.append(
            f"... {overflow} more rows in original at [[{original_ref}]]"
        )

    # Collect formula cells (all rows, all columns)
    for row in raw_rows:
        for cell in row:
            if _is_formula(cell):
                formula_count += 1
                col_letter = get_column_letter(cell.column)
                ref = f"{col_letter}{cell.row}"
                formula_val = str(cell.value)  # the formula string e.g. =SUM(A1:A5)
                cached = getattr(cell, "cached_value", None)
                if cached is not None:
                    formula_lines.append(f"{ref}: {formula_val}  → {cached}")
                else:
                    formula_lines.append(f"{ref}: {formula_val}")

    return body_lines, formula_lines, total_data_rows, formula_count, 1


# ──────────────────────────────────────────────────────────────────────────────
# Main handler
# ──────────────────────────────────────────────────────────────────────────────

def handle_xlsx(
    source_path: Path,
    *,
    pipeline_computed: dict,
    pipeline_ns: dict,
    original_ref: Optional[str] = None,
) -> HandlerResult:
    """Ingest an XLSX file.

    Parameters
    ----------
    source_path : Path
        Path to the .xlsx file.
    pipeline_computed : dict
        Top-level fields the pipeline computed.
    pipeline_ns : dict
        The pipeline: namespace dict.
    original_ref : str | None
        Vault-relative path for the overflow footer wikilink.
        Default: ``99 Workspace/_originals/<source_path.name>``

    Output structure
    ----------------
    ## Sheet: <name>

    <markdown table, ≤1000 data rows>

    ... N more rows in original at [[<original_ref>]]   ← only if overflow

    ## Formulas

    <cell>: <formula>  [→ <cached_value>]               ← every formula cell
    """
    if not _HAS_OPENPYXL:
        return HandlerResult(
            success=False,
            quarantine=True,
            quarantine_reason="missing_dependency_openpyxl",
            warnings=["openpyxl not installed"],
        )

    warnings: list[str] = []

    if original_ref is None:
        original_ref = f"99 Workspace/_originals/{source_path.name}"

    # ── Open workbook ──────────────────────────────────────────────────────
    # data_only=False: we want formula strings, not only cached values.
    # Keep formulas in the table; cached values shown in ## Formulas section.
    try:
        wb = openpyxl.load_workbook(str(source_path), data_only=False)
    except Exception as exc:
        return HandlerResult(
            success=False,
            quarantine=True,
            quarantine_reason="xlsx_open_error",
            warnings=[f"xlsx_open_error: {exc}"],
        )

    sheet_names = wb.sheetnames
    sheet_count = len(sheet_names)

    # ── Build body ─────────────────────────────────────────────────────────
    body_sections: list[str] = []
    total_formula_count = 0
    total_rows_across_sheets = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        body_sections.append(f"## Sheet: {sheet_name}")
        body_sections.append("")

        sheet_body, formula_lines, row_count, fml_count, tbl_count = _extract_sheet(
            ws,
            original_ref=original_ref,
        )

        total_formula_count += fml_count
        total_rows_across_sheets += row_count

        if sheet_body:
            body_sections.extend(sheet_body)
        else:
            body_sections.append("*(empty sheet)*")

        # ## Formulas section for this sheet
        if formula_lines:
            body_sections.append("")
            body_sections.append("## Formulas")
            body_sections.append("")
            body_sections.extend(formula_lines)

        body_sections.append("")
        body_sections.append("---")
        body_sections.append("")

    # Remove trailing separator
    while body_sections and body_sections[-1] in ("", "---"):
        body_sections.pop()

    body = "\n".join(body_sections)

    # ── Frontmatter ────────────────────────────────────────────────────────
    top: dict = {k: v for k, v in pipeline_computed.items()}
    top["file_type"] = "xlsx"
    top.setdefault("type", "spreadsheet")
    top["page_count"] = None  # not applicable to spreadsheets

    # ── Pipeline namespace ─────────────────────────────────────────────────
    from .text import compute_text_sha256   # CH-02
    pipeline_ns["provenance"] = _PROVENANCE_VALUE
    pipeline_ns["extractor"] = f"openpyxl@{_openpyxl_version()}"
    pipeline_ns["sheet_count"] = sheet_count
    pipeline_ns["sheet_names"] = sheet_names
    pipeline_ns["formula_count"] = total_formula_count
    pipeline_ns["row_count_total"] = total_rows_across_sheets
    pipeline_ns["row_cap"] = ROW_CAP
    pipeline_ns["text_sha256"] = compute_text_sha256(body)   # CH-02
    pipeline_ns["warnings"] = warnings

    return HandlerResult(
        success=True,
        body=body,
        frontmatter_top=top,
        pipeline_namespace=pipeline_ns,
        warnings=warnings,
    )
